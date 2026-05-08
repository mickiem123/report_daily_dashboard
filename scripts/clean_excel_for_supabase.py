from __future__ import annotations

import csv
import datetime as dt
import sys
from pathlib import Path

from openpyxl import load_workbook


OUTPUT_COLUMNS = [
    "ngay",
    "thi_phan_co_so",
    "thi_phan_cn",
    "thi_phan_ds",
    "gtgd_cs_ssi",
    "thanh_khoan_ttcs",
    "tong_du_no_margin",
    "slkh_margin",
    "du_no_t7",
    "slkh_t7",
    "du_no_trading_plus",
    "slkh_trading_plus",
    "slkh_register_mplus",
    "slkh_active_mplus",
    "du_no_mplus",
    "slkh_co_du_no_mplus",
    "giai_ngan_mplus",
    "slkh_giai_ngan_mplus",
    "du_no_spv",
    "ty_trong_spv",
    "thi_phan_phai_sinh",
    "thanh_khoan_tt_ps",
    "slkh_ps",
    "ty_le_slhd_dplus",
    "slkh_dplus",
    "ty_le_slkh_dplus",
    "kh_cancel_dplus",
    "kh_register_dplus",
    "kh_giu_qua_dem",
    "kh_sd_dplus",
    "slhd_giu_qua_dem",
    "du_no_dplus_giai_ngan",
    "du_no_dplus_cuoi_ngay",
    "so_du_scash",
    "so_du_casa_scash",
    "ty_le_scash_casa",
    "slkh_scash",
    "so_du_sfund",
    "slkh_sfund",
    "slkh_mo_moi",
]

# Excel columns (1-based) that are percentages in source files.
PERCENT_COLUMN_INDEXES = {2, 3, 4, 21, 22, 27, 29, 39}
ERROR_TOKENS = {"#DIV/0!", "#REF!", "#N/A", "#VALUE!", "#NAME?", "#NULL!", "#NUM!"}

# Excel columns (1-based) that are deprecated and intentionally dropped.
DROP_EXCEL_COLUMN_INDEXES = {19, 24, 26}


def parse_date(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        # Excel serial date system (1900-date system)
        if value <= 0:
            return None
        base = dt.datetime(1899, 12, 30)
        date_value = base + dt.timedelta(days=float(value))
        return date_value.date().isoformat()

    raw = str(value).strip()
    if not raw:
        return None
    raw = raw.replace("\\", "/").replace("-", "/")
    for fmt in ("%d/%b/%Y", "%d/%b/%y", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return dt.datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            continue
    for fmt in ("%Y/%m/%d",):
        try:
            return dt.datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def parse_numeric(value: object, excel_col_idx: int) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        num = float(value)
        if excel_col_idx in PERCENT_COLUMN_INDEXES and num > 1:
            num = num / 100.0
        return num

    raw = str(value).strip()
    if not raw:
        return None
    if raw.upper() in ERROR_TOKENS:
        return None

    is_percent_text = "%" in raw
    cleaned = raw.replace("%", "").replace(",", "").replace(" ", "")
    if not cleaned:
        return None

    try:
        num = float(cleaned)
    except ValueError:
        return None

    if is_percent_text or (excel_col_idx in PERCENT_COLUMN_INDEXES and num > 1):
        num = num / 100.0
    return num


def clean_excel(
    input_path: Path,
    output_path: Path,
    sheet_name: str = "data",
    fill_missing_numeric: float = 0.0,
) -> tuple[int, int]:
    wb = load_workbook(filename=input_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
    ws = wb[sheet_name]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    written = 0
    skipped = 0

    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(OUTPUT_COLUMNS)

        # Row 1 is headers in source files.
        for row in ws.iter_rows(min_row=2, max_col=43, values_only=True):
            ngay = parse_date(row[0] if len(row) > 0 else None)
            if ngay is None:
                skipped += 1
                continue

            cleaned_row = [ngay]
            for excel_col_idx in range(2, 44):
                if excel_col_idx in DROP_EXCEL_COLUMN_INDEXES:
                    continue
                value = row[excel_col_idx - 1] if len(row) >= excel_col_idx else None
                parsed = parse_numeric(value, excel_col_idx)
                cleaned_row.append(fill_missing_numeric if parsed is None else parsed)
            writer.writerow(cleaned_row)
            written += 1

    return written, skipped


def main() -> int:
    if len(sys.argv) < 3:
        print("Usage: python scripts/clean_excel_for_supabase.py <input.xlsx> <output.csv> [sheet_name]")
        return 1

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    sheet_name = sys.argv[3] if len(sys.argv) > 3 else "data"

    if not input_path.exists():
        print(f"Input file not found: {input_path}")
        return 2

    written, skipped = clean_excel(input_path, output_path, sheet_name=sheet_name)
    print(f"Wrote {written} rows to {output_path}")
    print(f"Skipped {skipped} rows (invalid/missing date)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
